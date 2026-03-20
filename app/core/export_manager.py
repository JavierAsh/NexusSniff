"""
export_manager.py — Exportación de capturas a CSV, JSON, Excel y PCAP.

Formatos soportados:
  - PCAP (.pcap): Binario estándar compatible con Wireshark/tshark
  - Excel (.xlsx): Hojas separadas de resumen, paquetes y protocolos
  - CSV (.csv): Tabla plana de paquetes
  - JSON (.json): Datos raw completos
"""

import csv
import json
import struct
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from PyQt6.QtWidgets import QFileDialog, QWidget, QMessageBox


# Paleta de colores por protocolo para Excel (formato ARGB hex, sin #)
PROTOCOL_EXCEL_COLORS = {
    'TCP':     ('FF0db9f2', 'FF001820'),   # cyan
    'UDP':     ('FF10b981', 'FF001a12'),   # verde
    'ICMP':    ('FFf59e0b', 'FF1a0e00'),   # amber
    'DNS':     ('FFf97316', 'FF1a0800'),   # naranja
    'HTTP':    ('FF8b5cf6', 'FF0d0020'),   # violeta
    'HTTPS':   ('FF06b6d4', 'FF001a1c'),   # cyan claro
    'SSH':     ('FFec4899', 'FF1a001f'),   # rosa
    'ARP':     ('FFec4899', 'FF1a001f'),   # rosa
    'FTP':     ('FF84cc16', 'FF0e1a00'),   # lima
    'Unknown': ('FF64748b', 'FF101418'),   # gris
}


# ─────────────────────────────────────────────────────────────────────────────
# PCAP Global Header constants
# Magic number 0xa1b2c3d4 → timestamps en microsegundos, byte-order nativo
# ─────────────────────────────────────────────────────────────────────────────
_PCAP_MAGIC          = 0xa1b2c3d4
_PCAP_VERSION_MAJOR  = 2
_PCAP_VERSION_MINOR  = 4
_PCAP_THISZONE       = 0          # GMT
_PCAP_SIGFIGS        = 0          # timestamp accuracy
_PCAP_SNAPLEN        = 65535      # max bytes per packet
_PCAP_LINKTYPE_ETHER = 1          # Ethernet (DLT_EN10MB)


class ExportManager:
    """Gestor de exportación de paquetes capturados."""

    @staticmethod
    def export_pcap(packets: List[Dict[str, Any]], parent: QWidget = None) -> bool:
        """Exporta paquetes al formato PCAP estándar (compatible Wireshark o tshark).

        El formato interno de salida incluye:
          - Global header (24 bytes): Magic number, versión, snaplen.
          - Por cada paquete: Record header (16 bytes) + datos crudos originales.

        No requiere dependencias externas; utiliza `struct.pack` de biblioteca estándar.

        Args:
            packets (list): Lista de diccionarios de paquetes capturados a exportar.
            parent (QWidget, optional): Instancia del widget padre para anclar el UI de diálogo. Defaults to None.

        Returns:
            bool: `True` si la exportación finalizó correctamente, `False` si el 
            usuario canceló o si hubo un error de escritura.
        """
        filepath, _ = QFileDialog.getSaveFileName(
            parent, "Exportar captura como PCAP",
            f"captura_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pcap",
            "PCAP (*.pcap);;Todos los archivos (*)"
        )
        if not filepath:
            return False

        # Contar paquetes con datos raw disponibles
        exportable = [
            p for p in packets
            if isinstance(p.get('raw_data'), (bytes, bytearray)) and p['raw_data']
        ]

        if not exportable:
            QMessageBox.warning(
                parent,
                "Sin datos raw",
                "No hay datos raw de paquetes para exportar en formato PCAP.\n"
                "Los datos raw solo están disponibles durante una captura activa."
            )
            return False

        try:
            with open(filepath, 'wb') as f:
                # ── Global Header ──
                # Format: magic(4) major(2) minor(2) thiszone(4) sigfigs(4) snaplen(4) linktype(4)
                global_header = struct.pack(
                    '<IHHiIII',
                    _PCAP_MAGIC,
                    _PCAP_VERSION_MAJOR,
                    _PCAP_VERSION_MINOR,
                    _PCAP_THISZONE,
                    _PCAP_SIGFIGS,
                    _PCAP_SNAPLEN,
                    _PCAP_LINKTYPE_ETHER,
                )
                f.write(global_header)

                # ── Packet Records ──
                for pkt in exportable:
                    raw = bytes(pkt['raw_data'])
                    ts = pkt.get('timestamp', 0.0)

                    # Descomponer timestamp en segundos y microsegundos
                    ts_sec  = int(ts)
                    ts_usec = int((ts - ts_sec) * 1_000_000)

                    incl_len = len(raw)                      # bytes capturados
                    orig_len = pkt.get('length', incl_len)   # longitud original en la red

                    # Limitar a snaplen
                    if incl_len > _PCAP_SNAPLEN:
                        raw = raw[:_PCAP_SNAPLEN]
                        incl_len = _PCAP_SNAPLEN

                    # Per-packet header: ts_sec(4) ts_usec(4) incl_len(4) orig_len(4)
                    pkt_header = struct.pack(
                        '<IIII',
                        ts_sec,
                        ts_usec,
                        incl_len,
                        orig_len,
                    )
                    f.write(pkt_header)
                    f.write(raw)

            return True

        except OSError as e:
            QMessageBox.critical(
                parent, "Error al exportar",
                f"No se pudo escribir el archivo PCAP:\n{e}"
            )
            return False

    @staticmethod
    def export_csv(packets: List[Dict[str, Any]], parent: QWidget = None) -> bool:
        """Exporta una vista en tabla plana de los paquetes a formato CSV demarcado por comas.
        
        Args:
            packets (list): Lista de paquetes capturados a exportar.
            parent (QWidget, optional): Widget base para el QFileDialog. Defaults to None.
            
        Returns:
            bool: `True` en caso de éxito, `False` en caso contrario.
        """
        filepath, _ = QFileDialog.getSaveFileName(
            parent, "Exportar captura como CSV",
            f"captura_nexussniff_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV (*.csv);;Todos los archivos (*)"
        )
        if not filepath:
            return False

        fields = ['number', 'timestamp', 'src_ip', 'dst_ip',
                  'src_port', 'dst_port', 'protocol', 'length', 'info']

        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
                writer.writeheader()
                for pkt in packets:
                    writer.writerow(pkt)
        except OSError as e:
            QMessageBox.critical(
                parent, "Error al exportar",
                f"No se pudo escribir el archivo CSV:\n{e}"
            )
            return False

        return True

    @staticmethod
    def export_json(packets: List[Dict[str, Any]], parent: QWidget = None) -> bool:
        """Convierte y guarda la captura entera a un archivo JSON.
        
        Automáticamente filtra los datos binarios crudos (`raw_data`) para prevenir 
        errores de serialización durante el proceso `json.dump`.
        
        Args:
            packets (list): Lista tabular de datos de paquetes.
            parent (QWidget, optional): Widget padre para el UI. Defaults to None.
            
        Returns:
            bool: Indicador booleano de éxito de la operación.
        """
        filepath, _ = QFileDialog.getSaveFileName(
            parent, "Exportar captura como JSON",
            f"captura_nexussniff_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "JSON (*.json);;Todos los archivos (*)"
        )
        if not filepath:
            return False

        clean_packets = [
            {k: v for k, v in pkt.items() if k != 'raw_data'}
            for pkt in packets
        ]

        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(clean_packets, f, indent=2, ensure_ascii=False, default=str)
        except OSError as e:
            QMessageBox.critical(
                parent, "Error al exportar",
                f"No se pudo escribir el archivo JSON:\n{e}"
            )
            return False

        return True

    @staticmethod
    def export_excel(
        packets: List[Dict[str, Any]],
        stats: Dict[str, Any] = None,
        parent: QWidget = None
    ) -> bool:
        """Exporta la captura local a un libro de Excel (.xlsx) estructurado en pestañas.
        
        Crea 3 hojas dedicadas:
        1. Resumen: Tablas de meta-estadísticas (paquetes por sec, distribución, etc).
        2. Paquetes: Lista de paquetes capturados estilizados con colores por protocolo.
        3. Protocolos: Análisis y gráficos porcentuales de la distribución.
        
        Requiere la librería de terceros `openpyxl`.
        
        Args:
            packets (list): Paquetes para exportar e iterar.
            stats (dict, optional): Estadística global proporcionada por el engine de captura. Defaults to None.
            parent (QWidget, optional): Instancia interfaz padre para mensajes. Defaults to None.
            
        Returns:
            bool: `True` tras un guardado de Excel satisfactorio, o `False` tras error / cancelación.
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                PatternFill, Font, Alignment, Border, Side, GradientFill
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                parent,
                "Dependencia faltante",
                "El módulo 'openpyxl' no está instalado.\n\n"
                "Instálalo manualmente con:\n  pip install openpyxl"
            )
            return False

        filepath, _ = QFileDialog.getSaveFileName(
            parent, "Exportar captura como Excel",
            f"captura_nexussniff_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel (*.xlsx);;Todos los archivos (*)"
        )
        if not filepath:
            return False

        wb = openpyxl.Workbook()

        # ── Estilos reutilizables ──
        def make_fill(hex_color: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex_color)

        def thin_border() -> Border:
            s = Side(style='thin', color='FFe2e8f0')
            return Border(left=s, right=s, top=s, bottom=s)

        header_font     = Font(name="Segoe UI", bold=True, color="FFFFFFFF", size=10)
        header_fill     = make_fill("FF0a0e18")
        header_align    = Alignment(horizontal='center', vertical='center', wrap_text=True)

        sub_header_font = Font(name="Segoe UI", bold=True, color="FF0db9f2", size=10)
        sub_header_fill = make_fill("FF161b2e")

        normal_font     = Font(name="Consolas", size=9, color="FF1e293b")
        mono_font       = Font(name="Consolas", size=9)
        label_font      = Font(name="Segoe UI", bold=True, size=10, color="FF334155")
        value_font      = Font(name="Segoe UI", size=10, color="FF1e293b")

        border = thin_border()
        center_align = Alignment(horizontal='center', vertical='center')
        left_align   = Alignment(horizontal='left',   vertical='center')

        # ════════════════════════════════════════
        # HOJA 1: RESUMEN
        # ════════════════════════════════════════
        ws_summary = wb.active
        ws_summary.title = "📋 Resumen"

        ws_summary.column_dimensions['A'].width = 28
        ws_summary.column_dimensions['B'].width = 40

        # Título
        ws_summary.merge_cells('A1:B1')
        title_cell = ws_summary['A1']
        title_cell.value = "NexusSniff — Resumen de Captura"
        title_cell.font = Font(name="Segoe UI", bold=True, size=14, color="FF0db9f2")
        title_cell.fill = make_fill("FF0a0e18")
        title_cell.alignment = center_align
        ws_summary.row_dimensions[1].height = 32

        ws_summary.merge_cells('A2:B2')
        ws_summary['A2'].value = ""
        ws_summary['A2'].fill = make_fill("FF0f1420")

        stats = stats or {}
        now = datetime.now()
        total_pkts   = len(packets)
        total_bytes  = sum(p.get('length', 0) for p in packets)

        # Calcular distribución de protocolos
        proto_dist: Dict[str, int] = {}
        for p in packets:
            proto = p.get('protocol', 'Unknown')
            proto_dist[proto] = proto_dist.get(proto, 0) + 1

        rows = [
            ("Fecha de exportación",   now.strftime("%Y-%m-%d %H:%M:%S")),
            ("Total de paquetes",      f"{total_pkts:,}"),
            ("Total de bytes",         f"{total_bytes:,} bytes"),
            ("Protocolos únicos",      str(len(proto_dist))),
            ("Paquetes por segundo",   f"{stats.get('packets_per_sec', 0):.1f} pkt/s"),
            ("Throughput",             f"{stats.get('bytes_per_sec', 0) / 1024:.1f} KB/s"),
            ("Paquetes perdidos",      str(stats.get('dropped_packets', 0))),
        ]

        for i, (label, value) in enumerate(rows, start=3):
            row = ws_summary.row_dimensions[i]
            row.height = 22

            lc = ws_summary.cell(row=i, column=1, value=label)
            lc.font  = label_font
            lc.fill  = make_fill("FFf1f5f9")
            lc.alignment = left_align
            lc.border = border

            vc = ws_summary.cell(row=i, column=2, value=value)
            vc.font  = value_font
            vc.fill  = make_fill("FFffffff")
            vc.alignment = left_align
            vc.border = border

        # ════════════════════════════════════════
        # HOJA 2: PAQUETES
        # ════════════════════════════════════════
        ws_packets = wb.create_sheet("📦 Paquetes")

        col_headers = [
            ("Nº",        8),
            ("Tiempo",    14),
            ("Origen",    22),
            ("Destino",   22),
            ("Protocolo", 12),
            ("Longitud",  10),
            ("Info",      50),
        ]
        col_keys = ['number', 'timestamp', 'src_ip', 'dst_ip', 'protocol', 'length', 'info']

        for col_idx, (header_text, width) in enumerate(col_headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws_packets.column_dimensions[col_letter].width = width
            cell = ws_packets.cell(row=1, column=col_idx, value=header_text)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border

        ws_packets.row_dimensions[1].height = 24
        ws_packets.freeze_panes = 'A2'

        # Estilos pre-creados para el loop (evita crear objetos por fila)
        proto_cell_font  = Font(name="Segoe UI", bold=True, size=9, color="FFFFFFFF")
        row_fill_even    = make_fill("FFffffff")
        row_fill_odd     = make_fill("FFf8fafc")
        proto_fill_cache: Dict[str, PatternFill] = {}

        for row_idx, pkt in enumerate(packets, start=2):
            proto = pkt.get('protocol', 'Unknown')
            colors = PROTOCOL_EXCEL_COLORS.get(proto, PROTOCOL_EXCEL_COLORS['Unknown'])
            fg_hex, _ = colors

            row_fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
            if fg_hex not in proto_fill_cache:
                proto_fill_cache[fg_hex] = make_fill(fg_hex)
            proto_fill = proto_fill_cache[fg_hex]

            # Construir valores
            src = pkt.get('src_ip', '')
            src_port = pkt.get('src_port', 0)
            if src_port:
                src = f"{src}:{src_port}"

            dst = pkt.get('dst_ip', '')
            dst_port = pkt.get('dst_port', 0)
            if dst_port:
                dst = f"{dst}:{dst_port}"

            ts = pkt.get('timestamp', 0)
            if isinstance(ts, float) and ts > 0:
                ts_str = f"{ts:.6f}"
            else:
                ts_str = str(ts)

            values = [
                pkt.get('number', row_idx - 1),
                ts_str,
                src,
                dst,
                proto,
                pkt.get('length', 0),
                pkt.get('info', ''),
            ]

            ws_packets.row_dimensions[row_idx].height = 18

            for col_idx, val in enumerate(values, start=1):
                cell = ws_packets.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = border
                cell.alignment = left_align

                if col_idx == 5:  # Protocolo
                    cell.font  = proto_cell_font
                    cell.fill  = proto_fill
                    cell.alignment = center_align
                else:
                    cell.font = mono_font
                    cell.fill = row_fill

        # ════════════════════════════════════════
        # HOJA 3: DISTRIBUCIÓN DE PROTOCOLOS
        # ════════════════════════════════════════
        ws_protos = wb.create_sheet("📊 Protocolos")

        ws_protos.column_dimensions['A'].width = 18
        ws_protos.column_dimensions['B'].width = 14
        ws_protos.column_dimensions['C'].width = 12

        prot_headers = [("Protocolo", 18), ("Paquetes", 14), ("Porcentaje", 12)]
        for col_idx, (hdr, _) in enumerate(prot_headers, start=1):
            cell = ws_protos.cell(row=1, column=col_idx, value=hdr)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border

        ws_protos.row_dimensions[1].height = 24
        ws_protos.freeze_panes = 'A2'

        total = sum(proto_dist.values()) or 1
        sorted_protos = sorted(proto_dist.items(), key=lambda x: x[1], reverse=True)

        proto_title_font = Font(name="Segoe UI", bold=True, size=10, color="FFFFFFFF")
        proto_data_font  = Font(name="Consolas", size=10)
        proto_data_fill  = make_fill("FFfafafa")

        for row_idx, (proto_name, count) in enumerate(sorted_protos, start=2):
            pct = count / total * 100
            colors = PROTOCOL_EXCEL_COLORS.get(proto_name, PROTOCOL_EXCEL_COLORS['Unknown'])
            fg_hex, _ = colors

            ws_protos.row_dimensions[row_idx].height = 20

            if fg_hex not in proto_fill_cache:
                proto_fill_cache[fg_hex] = make_fill(fg_hex)

            p_cell = ws_protos.cell(row=row_idx, column=1, value=proto_name)
            p_cell.font      = proto_title_font
            p_cell.fill      = proto_fill_cache[fg_hex]
            p_cell.alignment = center_align
            p_cell.border    = border

            c_cell = ws_protos.cell(row=row_idx, column=2, value=count)
            c_cell.font      = proto_data_font
            c_cell.alignment = center_align
            c_cell.border    = border
            c_cell.fill      = proto_data_fill

            pct_cell = ws_protos.cell(row=row_idx, column=3, value=f"{pct:.1f}%")
            pct_cell.font      = proto_data_font
            pct_cell.alignment = center_align
            pct_cell.border    = border
            pct_cell.fill      = proto_data_fill

        try:
            wb.save(filepath)
        except OSError as e:
            QMessageBox.critical(
                parent, "Error al exportar",
                f"No se pudo guardar el archivo Excel:\n{e}"
            )
            return False
        return True
